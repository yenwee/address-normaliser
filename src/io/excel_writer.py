"""Excel output: write results and colour-code rows by confidence.

Uses xlsxwriter for initial write (produces shared strings, editable in
Google Sheets/mobile Excel) and openpyxl for post-processing styling
(cell colours, wrap text, auto-width). After openpyxl save, the output
is post-processed to convert inline strings back to shared strings.
"""

import os
import re
import shutil
import tempfile
import zipfile
from xml.sax.saxutils import escape as xml_escape

import pandas as pd
from src.processing.mailability import inspect_mailing_block


def write_results(results: list[dict], output_path: str) -> None:
    """Write processed results to an Excel file.

    Args:
        results: List of result dicts with ICNO, NAME, MAILING_ADDRESS, CONFIDENCE.
        output_path: Path for the output Excel file.
    """
    out_df = pd.DataFrame(results, columns=["ICNO", "NAME", "MAILING_ADDRESS", "CONFIDENCE"])
    out_df["MAILING_ADDRESS"] = out_df["MAILING_ADDRESS"].fillna("")
    out_df.to_excel(output_path, index=False, engine="xlsxwriter")


def highlight_rows(path: str) -> None:
    """Colour-code Excel rows by confidence level.

    Red:    no address or no postcode (unmailable)
    Yellow: low/medium confidence (needs review)
    Green:  high confidence header row
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_header = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    wb = load_workbook(path)
    ws = wb.active

    # Header row
    for cell in ws[1]:
        cell.fill = green_header

    for row_idx in range(2, ws.max_row + 1):
        confidence = ws.cell(row=row_idx, column=4).value or 0
        address = str(ws.cell(row=row_idx, column=3).value or "")
        signals = inspect_mailing_block(address)
        has_postcode = signals["has_postcode"]
        has_street = signals["has_street"]
        has_house_number = signals["has_house_number"]

        if confidence == 0 or not address.strip():
            fill = red
        elif not has_postcode:
            fill = red
        elif not has_street:
            fill = red
        elif not has_house_number:
            fill = yellow
        elif confidence < 0.6:
            fill = yellow
        else:
            continue

        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).fill = fill

    # Enable wrap text on address column so multi-line addresses display properly
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).alignment = wrap

    # Auto-width columns
    for col_idx in range(1, 5):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row + 1, 50)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    wb.save(path)
    _convert_inline_to_shared_strings(path)


def _convert_inline_to_shared_strings(xlsx_path: str) -> None:
    """Rewrite xlsx to convert inlineStr cells to shared strings.

    openpyxl 3.x writes all string cells as <is><t>value</t></is> which
    Google Sheets and mobile Excel render as read-only. This converts
    them to shared string references, preserving cell formatting.
    """
    cell_re = re.compile(
        r'<c([^>]*?)\s+t="inlineStr"([^>]*?)>\s*<is>\s*<t(\s+[^>]*?)?>(.*?)</t>\s*</is>\s*</c>',
        re.DOTALL,
    )
    empty_cell_re = re.compile(r'<c([^>]*?)\s+t="inlineStr"([^>]*?)\s*/>')
    ct_pattern = re.compile(r'</Types>')

    shared_strings: dict[str, int] = {}

    def add_string(s: str) -> int:
        if s not in shared_strings:
            shared_strings[s] = len(shared_strings)
        return shared_strings[s]

    def unescape_xml(s: str) -> str:
        return (
            s.replace("&lt;", "<")
             .replace("&gt;", ">")
             .replace("&quot;", '"')
             .replace("&apos;", "'")
             .replace("&amp;", "&")
        )

    def replace_cell(match):
        before_t = match.group(1) or ""
        after_t = match.group(2) or ""
        text_content = unescape_xml(match.group(4) or "")
        idx = add_string(text_content)
        attrs = (before_t + after_t).strip()
        if attrs:
            return f'<c {attrs} t="s"><v>{idx}</v></c>'
        return f'<c t="s"><v>{idx}</v></c>'

    def replace_empty_cell(match):
        before_t = match.group(1) or ""
        after_t = match.group(2) or ""
        attrs = (before_t + after_t).strip()
        if attrs:
            return f'<c {attrs}/>'
        return '<c/>'

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    try:
        with zipfile.ZipFile(xlsx_path, "r") as src, zipfile.ZipFile(
            tmp_path, "w", zipfile.ZIP_DEFLATED
        ) as dst:
            sheet_names = [
                n for n in src.namelist() if n.startswith("xl/worksheets/") and n.endswith(".xml")
            ]
            has_shared_strings = "xl/sharedStrings.xml" in src.namelist()

            for name in src.namelist():
                data = src.read(name)
                if name in sheet_names:
                    xml_str = data.decode("utf-8")
                    new_xml = cell_re.sub(replace_cell, xml_str)
                    new_xml = empty_cell_re.sub(replace_empty_cell, new_xml)
                    data = new_xml.encode("utf-8")
                elif name == "[Content_Types].xml" and not has_shared_strings:
                    ct_str = data.decode("utf-8")
                    ss_override = (
                        '<Override PartName="/xl/sharedStrings.xml" '
                        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                    )
                    if "sharedStrings" not in ct_str:
                        ct_str = ct_pattern.sub(ss_override + "</Types>", ct_str)
                    data = ct_str.encode("utf-8")
                elif name == "xl/_rels/workbook.xml.rels" and not has_shared_strings:
                    rels_str = data.decode("utf-8")
                    if "sharedStrings" not in rels_str:
                        import uuid
                        rid = f"rId{uuid.uuid4().hex[:8]}"
                        rel = (
                            f'<Relationship Id="{rid}" '
                            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                            'Target="sharedStrings.xml"/>'
                        )
                        rels_str = rels_str.replace("</Relationships>", rel + "</Relationships>")
                    data = rels_str.encode("utf-8")
                dst.writestr(name, data)

            count = len(shared_strings)
            if count > 0:
                sorted_strings = sorted(shared_strings.items(), key=lambda x: x[1])
                parts = [
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n',
                    f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="{count}" uniqueCount="{count}">',
                ]
                for s, _ in sorted_strings:
                    escaped = xml_escape(s)
                    preserve = ' xml:space="preserve"' if s != s.strip() or "\n" in s else ""
                    parts.append(f"<si><t{preserve}>{escaped}</t></si>")
                parts.append("</sst>")
                dst.writestr("xl/sharedStrings.xml", "".join(parts))

        shutil.move(tmp_path, xlsx_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
