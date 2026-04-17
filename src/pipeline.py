"""Main processing pipeline for Malaysian address normalisation.

Reads Excel files containing ICNO, NAME, and ADDR columns, normalises
addresses through parsing, clustering, scoring, validation, and optional
geocoding, then writes a clean output Excel with mailing blocks.
"""

import logging
import re
from collections import Counter
from pathlib import Path
from typing import Optional

import pandas as pd

from src.clusterer import cluster_addresses
from src.config import (
    ADDR_COLUMN_PREFIX,
    CONFIDENCE_THRESHOLD,
    IC_COLUMN,
    NAME_COLUMN,
    NOMINATIM_ENABLED,
)
from src.formatter import format_mailing_block
from src.nominatim import geocode_address
from src.normaliser import normalise_address, normalise_state
from src.parser import parse_all_addresses
from src.scorer import score_completeness
from src.validator import POSTCODE_STATE_PREFIXES, PostcodeValidator

logger = logging.getLogger(__name__)

MAX_COMPLETENESS_SCORE = 12
POSTCODES_PATH = "data/postcodes.json"

_HEADER_IC_VALUES = frozenset({"ic", "icno"})
_ADDR_COL_RE = re.compile(rf"^{ADDR_COLUMN_PREFIX}(\d+)$", re.IGNORECASE)


def _read_excel(path: str) -> pd.DataFrame:
    """Read an Excel file, supporting both .xls and .xlsx formats.

    For .xls files, uses xlrd with ignore_workbook_corruption=True.
    For .xlsx files, uses openpyxl.

    Args:
        path: Path to the Excel file.

    Returns:
        A pandas DataFrame with the file contents.
    """
    ext = Path(path).suffix.lower()

    if ext == ".xls":
        import xlrd

        workbook = xlrd.open_workbook(path, ignore_workbook_corruption=True)
        return pd.read_excel(workbook, engine="xlrd")

    return pd.read_excel(path, engine="openpyxl")


def _get_addr_columns(df: pd.DataFrame) -> list[str]:
    """Find and sort ADDR columns by their numeric suffix.

    Args:
        df: DataFrame whose columns to inspect.

    Returns:
        Sorted list of column names matching ADDR<number> pattern.
    """
    addr_cols = []
    for col in df.columns:
        match = _ADDR_COL_RE.match(str(col))
        if match:
            addr_cols.append((int(match.group(1)), str(col)))

    addr_cols.sort(key=lambda pair: pair[0])
    return [col for _, col in addr_cols]


_POPULARITY_WEIGHT = 0.5


def _select_best_address(
    clusters: list[list[dict]],
    raw_postcode_counts: Optional[Counter] = None,
) -> tuple[Optional[dict], float]:
    """Select the best address from clustered address variants.

    Scores each cluster by: len(cluster) * max(score_completeness(addr)),
    factored by postcode consistency and raw postcode popularity.

    Args:
        clusters: List of address clusters from cluster_addresses.
        raw_postcode_counts: Postcode frequency across ALL raw ADDR columns.

    Returns:
        Tuple of (best_address_dict, confidence) or (None, 0.0).
    """
    if not clusters:
        return None, 0.0

    max_raw = max(raw_postcode_counts.values()) if raw_postcode_counts else 0

    best_cluster = None
    best_cluster_score = -1

    for cluster in clusters:
        if not cluster:
            continue
        max_addr_score = max(score_completeness(addr) for addr in cluster)
        # Factor in postcode consistency: clusters where members agree on postcode score higher
        postcodes = [a["postcode"] for a in cluster if a.get("postcode", "").strip()]
        if postcodes:
            postcode_counts = Counter(postcodes)
            postcode_consistency = max(postcode_counts.values()) / len(cluster)
            dominant_pc = postcode_counts.most_common(1)[0][0]
        else:
            postcode_consistency = 0
            dominant_pc = ""
        cluster_score = len(cluster) * max_addr_score * (0.5 + 0.5 * postcode_consistency)
        # Factor in raw postcode popularity across all ADDR columns
        if raw_postcode_counts and max_raw and dominant_pc:
            popularity = raw_postcode_counts.get(dominant_pc, 0) / max_raw
            cluster_score *= (1 - _POPULARITY_WEIGHT + _POPULARITY_WEIGHT * popularity)
        if cluster_score > best_cluster_score:
            best_cluster_score = cluster_score
            best_cluster = cluster

    if best_cluster is None:
        return None, 0.0

    best_addr = _select_from_cluster(best_cluster)
    confidence = min(score_completeness(best_addr) / MAX_COMPLETENESS_SCORE, 1.0)

    return best_addr, confidence


def _select_from_cluster(cluster: list[dict]) -> dict:
    """Select the best address from a cluster using completeness + consensus.

    When multiple addresses have similar completeness scores, prefer the one
    whose street name pattern is agreed upon by more cluster members.
    E.g. if 2 addresses say 'LORONG PUTERI GUNUNG' and 1 says 'LORONG 3',
    prefer the descriptive version even if the generic one scores slightly higher.
    """
    if len(cluster) == 1:
        return cluster[0]

    scored = [(addr, score_completeness(addr)) for addr in cluster]
    max_score = max(s for _, s in scored)

    # If clear winner (3+ points ahead of all others), just pick it
    runner_up = sorted(set(s for _, s in scored), reverse=True)
    if len(runner_up) >= 2 and runner_up[0] - runner_up[1] >= 3:
        return max(cluster, key=score_completeness)

    # Close scores — use consensus tiebreaker
    # Within 2 points of max, factor in how many cluster members agree
    from rapidfuzz import fuzz
    top_addrs = [(a, s) for a, s in scored if s >= max_score - 2]

    best = None
    best_combined = -1
    for addr, sc in top_addrs:
        consensus = sum(
            1 for other in cluster
            if fuzz.token_sort_ratio(addr["address_line"], other["address_line"]) > 70
        )
        # Consensus can override up to a 2-point score gap
        combined = sc + consensus * 2
        if combined > best_combined:
            best_combined = combined
            best = addr

    return best


def _is_header_row(row: pd.Series) -> bool:
    """Check if a row is a repeated header (ICNO value is a header label)."""
    ic_value = row.get(IC_COLUMN, "")
    if pd.isna(ic_value):
        return False
    return str(ic_value).strip().lower() in _HEADER_IC_VALUES


def _strip_leaked_fields(addr: dict) -> dict:
    """Remove state names and postcode+city patterns that leaked into address lines."""
    import re
    from src.parser import KNOWN_STATES

    cleaned = dict(addr)
    state = cleaned.get("state", "").upper()
    postcode = cleaned.get("postcode", "")
    city = cleaned.get("city", "").upper()

    for key in ("address_line", "address_line2", "address_line3"):
        val = cleaned.get(key, "").strip()
        upper_val = val.upper()

        # Strip if entire field is just a state name
        if upper_val in KNOWN_STATES or upper_val == state:
            cleaned[key] = ""
            continue

        # Strip "postcode city" pattern from address lines (e.g. "70300 SEREMBAN")
        if postcode and city:
            pattern = f"{postcode}\\s*{re.escape(city)}"
            stripped = re.sub(pattern, "", upper_val, flags=re.IGNORECASE).strip()
            if stripped != upper_val:
                cleaned[key] = stripped
                continue

        # Strip standalone postcode from address lines (e.g. "... 06150 ALOR SETAR")
        if postcode:
            stripped = re.sub(rf"\b{postcode}\b", "", upper_val).strip()
            stripped = re.sub(r"\s+", " ", stripped)
            if stripped != upper_val:
                cleaned[key] = stripped
                val = cleaned[key]
                upper_val = val.upper()

        # Strip city name from END of address lines (e.g. "JLN ADABI KOTA BHARU" when city=KOTA BHARU)
        if city and len(city) > 3 and upper_val.endswith(city):
            stripped = upper_val[: -len(city)].strip()
            if stripped:
                cleaned[key] = stripped

        # Strip state name from END of address lines
        if state and len(state) > 2 and upper_val.endswith(state):
            stripped = upper_val[: -len(state)].strip()
            if stripped:
                cleaned[key] = stripped

    return cleaned


_MERGE_WORDS = frozenset({"JALAN", "KAMPUNG", "LORONG", "TAMAN", "BANDAR", "SUNGAI", "BATU", "DESA"})
_MERGE_WORDS_NEED_NUMBER = frozenset({"BATU"})


def _merge_standalone_words(addr: dict) -> dict:
    """Merge standalone generic words (JALAN, KAMPUNG etc.) with adjacent lines.

    Source data sometimes splits 'Kampung' or 'Jalan' into a separate comma field,
    producing standalone address lines like just 'JALAN' or 'KAMPUNG'.
    These should be merged with the next non-empty line.
    """
    cleaned = dict(addr)
    lines = [cleaned.get("address_line", ""), cleaned.get("address_line2", ""), cleaned.get("address_line3", "")]

    merged = []
    carry = ""
    for i, line in enumerate(lines):
        if carry:
            if line:
                # BATU should only merge when the following token starts with a digit
                # (e.g. "BATU 7"). Otherwise drop the stale label.
                if carry.upper() in _MERGE_WORDS_NEED_NUMBER:
                    first_tok = line.strip().split()[0] if line.strip() else ""
                    if not re.match(r"^\d", first_tok):
                        carry = ""
                        merged.append(line)
                        continue
                line = f"{carry} {line}".strip()
                carry = ""
            else:
                # No next line -- check remaining lines for something to merge with
                remaining = [l for l in lines[i + 1:] if l.strip()]
                if remaining:
                    # Will merge with next non-empty line in a future iteration
                    merged.append("")  # placeholder for this empty slot
                    continue
                # Nothing ahead -- append to previous if available, else drop number-needing words
                if carry.upper() in _MERGE_WORDS_NEED_NUMBER and not merged:
                    carry = ""
                    merged.append(line)
                    continue
                if merged:
                    merged[-1] = f"{merged[-1]} {carry}".strip() if merged[-1] else carry
                else:
                    merged.append(carry)
                carry = ""
                merged.append(line)
                continue
        if line.strip().upper() in _MERGE_WORDS:
            carry = line.strip()
        else:
            merged.append(line)

    if carry:
        if carry.upper() in _MERGE_WORDS_NEED_NUMBER and not merged:
            pass
        elif merged:
            merged[-1] = f"{merged[-1]} {carry}".strip() if merged[-1] else carry
        else:
            merged.append(carry)

    while len(merged) < 3:
        merged.append("")

    cleaned["address_line"] = merged[0]
    cleaned["address_line2"] = merged[1]
    cleaned["address_line3"] = merged[2]

    return cleaned


def _find_best_cluster(clusters: list) -> list:
    """Return the cluster that matches _select_best_address's scoring."""
    best = None
    best_score = -1
    for cluster in clusters:
        if not cluster:
            continue
        max_addr_score = max(score_completeness(a) for a in cluster)
        postcodes = [a["postcode"] for a in cluster if a.get("postcode", "").strip()]
        if postcodes:
            postcode_consistency = max(Counter(postcodes).values()) / len(cluster)
        else:
            postcode_consistency = 0
        score = len(cluster) * max_addr_score * (0.5 + 0.5 * postcode_consistency)
        if score > best_score:
            best_score = score
            best = cluster
    return best or []


def _enrich_from_cluster(best_addr: dict, clusters: list) -> dict:
    """Light merge: enrich the best address with specific info from cluster siblings.

    Currently handles:
    - Missing JALAN/LORONG prefix before street number patterns (e.g. "2/12A")
    - Cross-cluster street name borrowing when addresses describe same place
    """
    import re

    # Find the cluster that contains best_addr (not always the highest-scoring
    # cluster, since popularity scoring may have selected a different one)
    best_cluster = None
    for c in clusters:
        if any(a is best_addr for a in c):
            best_cluster = c
            break
    if best_cluster is None:
        best_cluster = _find_best_cluster(clusters)

    if len(best_cluster) < 2 and len(clusters) < 2:
        return best_addr

    enriched = dict(best_addr)
    addr_line = enriched.get("address_line", "")

    # Check for street number pattern without JALAN/LORONG prefix
    # e.g. "NO 69 2/12A" — the "2/12A" looks like a street ref
    street_num_pattern = re.search(r"\b(\d+/\d+\w?)\b", addr_line)
    if street_num_pattern:
        num = street_num_pattern.group(1)
        has_prefix = bool(re.search(
            rf"\b(?:JALAN|LORONG|PERSIARAN|LEBUH)\s+{re.escape(num)}",
            addr_line, re.IGNORECASE,
        ))

        if not has_prefix:
            # Search cluster siblings for JALAN/LORONG before this number
            for sibling in best_cluster:
                sib_line = sibling.get("address_line", "") + " " + sibling.get("address_line2", "")
                m = re.search(
                    rf"\b(JALAN|LORONG|PERSIARAN|LEBUH)\s+{re.escape(num)}",
                    sib_line, re.IGNORECASE,
                )
                if m:
                    prefix = m.group(1)
                    enriched["address_line"] = re.sub(
                        rf"\b{re.escape(num)}\b",
                        f"{prefix} {num}",
                        enriched["address_line"],
                        count=1,
                    )
                    break

    # Cross-cluster street enrichment: if selected address has no street keyword
    # on EITHER line, borrow from another cluster describing the SAME place.
    addr_line = enriched.get("address_line", "")
    addr_line2 = enriched.get("address_line2", "")
    _STREET_KW_RE = re.compile(r"\b(?:JALAN|LORONG|PERSIARAN|LEBUH)\b", re.IGNORECASE)
    if not _STREET_KW_RE.search(addr_line) and not _STREET_KW_RE.search(addr_line2):
        best_pc = enriched.get("postcode", "")
        if best_pc:
            from rapidfuzz import fuzz as _fuzz
            best_text = " ".join(filter(None, [
                enriched.get("address_line", ""),
                enriched.get("address_line2", ""),
                enriched.get("city", ""),
            ])).upper()

            for other_cluster in clusters:
                if other_cluster is best_cluster:
                    continue
                for sib in other_cluster:
                    sib_pc = sib.get("postcode", "")
                    if sib_pc != best_pc:
                        continue
                    sib_line = sib.get("address_line", "")
                    if not _STREET_KW_RE.search(sib_line):
                        continue
                    # Guard 1: reject if borrowed line embeds a postcode
                    if re.search(r"\b\d{5}\b", sib_line):
                        continue
                    # Guard 2: reject if current content not found in borrowed
                    addr_clean = re.sub(r"[\s\-.]", "", addr_line.upper())
                    sib_clean = re.sub(r"[\s\-.]", "", sib_line.upper())
                    if addr_clean and addr_clean[:5] not in sib_clean:
                        # Clusters must also be semantically similar
                        sib_text = " ".join(filter(None, [
                            sib.get("address_line", ""),
                            sib.get("address_line2", ""),
                            sib.get("city", ""),
                        ])).upper()
                        if _fuzz.token_sort_ratio(best_text, sib_text) < 55:
                            continue
                    enriched["address_line"] = sib_line
                    break
                else:
                    continue
                break

    return enriched


def _ensemble_enhance(best_addr: dict, cluster: list) -> dict:
    """Fill missing fields from cluster majority vote. Never replace non-empty fields."""
    import re
    from collections import Counter
    from src.parser import KNOWN_STATES

    if len(cluster) < 2:
        return best_addr

    enhanced = dict(best_addr)

    # Fill empty address_line2 from cluster siblings
    if not enhanced.get("address_line2", "").strip():
        l2_candidates = []
        for a in cluster:
            l2 = a.get("address_line2", "").strip()
            if l2 and l2.upper() not in KNOWN_STATES:
                if not re.search(r"\b\d{5}\b", l2):
                    l2_candidates.append(l2)
        if l2_candidates:
            enhanced["address_line2"] = Counter(l2_candidates).most_common(1)[0][0]

    # Fill empty postcode
    if not enhanced.get("postcode", "").strip():
        pcs = [a["postcode"] for a in cluster if re.match(r"^\d{5}$", a.get("postcode", ""))]
        if pcs:
            enhanced["postcode"] = Counter(pcs).most_common(1)[0][0]

    # Fill empty city
    if not enhanced.get("city", "").strip():
        cities = [a["city"] for a in cluster if a["city"].strip()]
        if cities:
            enhanced["city"] = Counter(cities).most_common(1)[0][0]

    # Fill empty state
    if not enhanced.get("state", "").strip():
        states = [a["state"] for a in cluster if a["state"].strip()]
        if states:
            enhanced["state"] = Counter(states).most_common(1)[0][0]

    # Word-level spelling correction: fix individual misspelled words using
    # cluster majority vote. Only swaps words at the same position when the
    # majority spelling differs. Requires 3+ structurally similar lines.
    addr_line = enhanced.get("address_line", "").strip()
    if addr_line and len(cluster) >= 3:
        from rapidfuzz import fuzz
        words = addr_line.split()
        similar_lines = []
        for a in cluster:
            sib = a.get("address_line", "").strip()
            if sib and len(sib.split()) == len(words):
                similar_lines.append(sib.split())

        if len(similar_lines) >= 2:
            corrected = list(words)
            for i, word in enumerate(words):
                if re.match(r"^\d+$", word) or len(word) <= 2:
                    continue
                spellings = Counter()
                for sib_words in similar_lines:
                    sib_word = sib_words[i]
                    if fuzz.ratio(word.upper(), sib_word.upper()) > 60:
                        spellings[sib_word] += 1
                if not spellings:
                    continue
                majority_word, majority_count = spellings.most_common(1)[0]
                current_count = spellings.get(word, 0)
                # Only swap if majority is larger, words are similar but different,
                # and we're not dropping a suffix (e.g. 1084D -> 1084)
                if (majority_count > current_count
                        and majority_word.upper() != word.upper()
                        and fuzz.ratio(word.upper(), majority_word.upper()) > 60
                        and not (len(majority_word) < len(word)
                                 and word.startswith(majority_word))):
                    corrected[i] = majority_word
            enhanced["address_line"] = " ".join(corrected)

    return enhanced


def _apply_geocode_fallback(addr: dict) -> dict:
    """Use Nominatim geocoding to fill missing fields on low-confidence addresses.

    Args:
        addr: The address dict to enrich.

    Returns:
        A copy of addr with any missing fields filled from geocode results.
    """
    query = format_mailing_block(addr)
    result = geocode_address(query)
    if result is None:
        return addr

    enriched = dict(addr)

    if not enriched.get("postcode") and result.get("postcode"):
        enriched["postcode"] = result["postcode"]
    if not enriched.get("city") and result.get("city"):
        enriched["city"] = result["city"].upper()
    if not enriched.get("state") and result.get("state"):
        enriched["state"] = result["state"].upper()
    if not enriched.get("address_line") and result.get("road"):
        enriched["address_line"] = result["road"].upper()

    return enriched


def process_file(input_path: str, output_path: str) -> dict:
    """Process an Excel file of Malaysian addresses through the full pipeline.

    Flow per row:
        1. Parse all ADDR columns into structured address dicts
        2. Normalise each parsed address
        3. Cluster normalised addresses by similarity
        4. Select the best address from the best cluster
        5. Validate and correct postcode/city/state
        6. Optionally geocode low-confidence addresses
        7. Format as a mailing block

    Args:
        input_path: Path to the input Excel file.
        output_path: Path for the output Excel file.

    Returns:
        Stats dict with keys: total, processed, low_confidence, no_address.
    """
    df = _read_excel(input_path)
    addr_columns = _get_addr_columns(df)

    validator = PostcodeValidator(POSTCODES_PATH)

    stats = {
        "total": 0,
        "processed": 0,
        "low_confidence": 0,
        "no_address": 0,
    }

    results = []

    for _, row in df.iterrows():
        if _is_header_row(row):
            continue

        stats["total"] += 1

        ic = str(row.get(IC_COLUMN, "")).strip()
        name = str(row.get(NAME_COLUMN, "")).strip()

        parsed_addresses = parse_all_addresses(row, addr_columns)

        if not parsed_addresses:
            stats["no_address"] += 1
            stats["processed"] += 1
            results.append({
                "ICNO": ic,
                "NAME": name,
                "MAILING_ADDRESS": "",
                "CONFIDENCE": 0.0,
            })
            continue

        normalised = [normalise_address(addr) for addr in parsed_addresses]

        clusters = cluster_addresses(normalised)

        # Count raw postcode frequency across all ADDR columns for popularity scoring
        raw_pc_counts = Counter()
        for col in addr_columns:
            if col not in row.index:
                continue
            val = str(row[col]) if pd.notna(row[col]) else ""
            pc_match = re.search(r"\b(\d{5})\b", val)
            if pc_match:
                raw_pc_counts[pc_match.group(1)] += 1

        best_addr, confidence = _select_best_address(clusters, raw_pc_counts)

        if best_addr is None:
            stats["no_address"] += 1
            stats["processed"] += 1
            results.append({
                "ICNO": ic,
                "NAME": name,
                "MAILING_ADDRESS": "",
                "CONFIDENCE": 0.0,
            })
            continue

        # Light merge: enrich best address with info from cluster siblings
        best_addr = _enrich_from_cluster(best_addr, clusters)

        # Ensemble: fill missing fields from cluster majority vote
        best_cluster = _find_best_cluster(clusters)
        best_addr = _ensemble_enhance(best_addr, best_cluster)

        corrected, _ = validator.correct_address(best_addr)

        # Re-normalise state after validator (DB may return "WP KUALA LUMPUR")
        corrected["state"] = normalise_state(corrected.get("state", ""))

        # Fill missing state from postcode prefix as fallback
        if not corrected.get("state", "").strip() and corrected.get("postcode", "").strip():
            prefix = corrected["postcode"][:2]
            inferred = POSTCODE_STATE_PREFIXES.get(prefix, "")
            if inferred:
                corrected["state"] = normalise_state(inferred)

        # Clean up address lines
        corrected = _strip_leaked_fields(corrected)
        corrected = _merge_standalone_words(corrected)

        confidence = min(
            score_completeness(corrected) / MAX_COMPLETENESS_SCORE, 1.0
        )

        if NOMINATIM_ENABLED and confidence < CONFIDENCE_THRESHOLD:
            corrected = _apply_geocode_fallback(corrected)
            confidence = min(
                score_completeness(corrected) / MAX_COMPLETENESS_SCORE, 1.0
            )

        if confidence < CONFIDENCE_THRESHOLD:
            stats["low_confidence"] += 1

        mailing = format_mailing_block(corrected)

        stats["processed"] += 1
        results.append({
            "ICNO": ic,
            "NAME": name,
            "MAILING_ADDRESS": mailing,
            "CONFIDENCE": round(confidence, 2),
        })

    out_df = pd.DataFrame(results, columns=["ICNO", "NAME", "MAILING_ADDRESS", "CONFIDENCE"])
    out_df["MAILING_ADDRESS"] = out_df["MAILING_ADDRESS"].fillna("")
    out_df.to_excel(output_path, index=False, engine="openpyxl")

    _highlight_rows(output_path)

    logger.info(
        "Pipeline complete: %d total, %d processed, %d low confidence, %d no address",
        stats["total"],
        stats["processed"],
        stats["low_confidence"],
        stats["no_address"],
    )

    return stats


def _highlight_rows(path: str) -> None:
    """Colour-code Excel rows by confidence level.

    Red:    no address or no postcode (unmailable)
    Yellow: low/medium confidence (needs review)
    Green:  high confidence header row
    """
    import re
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, PatternFill

    from src.parser import KNOWN_STATES
    from src.normaliser import STATE_MAPPING
    known_states_upper = KNOWN_STATES | {v.upper() for v in STATE_MAPPING.values()}

    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_header = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    wb = load_workbook(path)
    ws = wb.active

    # Header row
    for cell in ws[1]:
        cell.fill = green_header

    for row_idx in range(2, ws.max_row + 1):
        confidence = ws.cell(row=row_idx, column=4).value or 0
        address = str(ws.cell(row=row_idx, column=3).value or "")
        has_postcode = bool(re.search(r"\b\d{5}\b", address))

        # Check if address has a street/house component (not just postcode+city+state)
        lines = [l.strip() for l in address.split("\n") if l.strip()]
        has_street = any(
            not re.match(r"^\d{5}\s", l) and l.upper() not in known_states_upper
            for l in lines
        )

        # Check for house/lot number keyword (NO, LOT, UNIT, BLK, or standalone leading digit on first line)
        addr_lines = [l.strip() for l in address.split("\n") if l.strip()]
        street_lines = [l for l in addr_lines
                        if not re.match(r"^\d{5}\s", l) and l.upper() not in known_states_upper]
        street_text = " ".join(street_lines)
        # Simple check: does the street text contain any number?
        # Addresses with no number at all (just area/village names) are incomplete.
        has_house_number = bool(re.search(r"\d", street_text))

        if confidence == 0 or not address.strip():
            fill = red
        elif not has_postcode:
            fill = red
        elif not has_street:
            fill = red
        elif not has_house_number:
            fill = yellow
        elif confidence < 0.6:
            fill = yellow
        else:
            continue

        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).fill = fill

    # Enable wrap text on address column so multi-line addresses display properly
    wrap = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).alignment = wrap

    # Auto-width columns
    for col_idx in range(1, 5):
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, min(ws.max_row + 1, 50)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    wb.save(path)
